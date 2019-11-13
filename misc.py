import time
from openpyxl import load_workbook, Workbook


def getLocalHour():
    secEpoch = time.time()
    localTime = time.localtime(secEpoch)
    return localTime.tm_hour

def getLocalDate():
    secEpoch = time.time()
    localTime = time.localtime(secEpoch)
    localDay = localTime.tm_mday
    localMonth = localTime.tm_mon
    localYear = localTime.tm_year

    return (localDay, localMonth, localYear)

def getNextIndex(filename, sheet):
    sheet = workbook.active
    max_row = sheet.max_row
    return max_row + 1


def excelWriter(date, weight, filename):
    print("Writing")
    workbook = load_workbook(filename = filename)
    sheet = workbook.active
    index = getNextIndex(filename, sheet)
    
    dateCell = "A" + str(index)
    weightCell = "B" + str(index)

    dateFormated = str(date[0]) + "." + str(date[1]) + "." + str(date[2])


    sheet[dateCell] = dateFormated
    sheet[weightCell] = weight
    workbook.save(filename=filename)

def measureWeight():
    # Measure weight
    return 0


def logg(filename):
    print('Logging initiated')
    weight = measureWeight()

    date = getLocalDate()
    excelWriter(date,weight,filename)


def calibrate():
    print('Calibration initiated')
